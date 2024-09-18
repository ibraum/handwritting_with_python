import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Exemple de données
data = {
    'Question': ['What is your name?', 'What is your address?', 'Describe your issue in detail.'],
    'Answer': ['John Doe', '1234 Main St, Anytown, USA', 'The issue started when I tried to update the software. It caused several bugs in the system.']
}

# Créer un DataFrame
df = pd.DataFrame(data)

# Sauvegarder le DataFrame en Excel
df.to_excel('output.xlsx', index=False)

# Charger le fichier Excel avec openpyxl pour faire des modifications supplémentaires
wb = load_workbook('output.xlsx')
ws = wb.active

# Appliquer le wrap text et ajuster la hauteur des lignes
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Ajuster la largeur des colonnes
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # obtenir la lettre de la colonne
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Sauvegarder les modifications
wb.save('output_formatted.xlsx')
